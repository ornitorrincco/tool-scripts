from openpyxl import load_workbook

data = load_workbook('test.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1'].value
# sheet.cell(row=1, column=2).value
